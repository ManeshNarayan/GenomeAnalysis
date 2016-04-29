import time
import math
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, WebDriverException
from selenium.webdriver.common.keys import Keys

wb=openpyxl.load_workbook('Freqdat.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
col = 1
while (col<60):
	for i in range(3, 350):
		k=i+1 
		for j in range(i+1, 352):
			if sheet.cell(row=j,column=col).value != ' ':	
				temp1 = sheet.cell(row=j,column=col).value
				temp2 = sheet.cell(row=j,column=col+1).value
				print(temp2)
				sheet.cell(row=j,column=col).value = ' '
				sheet.cell(row=j,column=col+1).value = ' '
				sheet.cell(row=k,column=col).value = temp1
				sheet.cell(row=k,column=col+1).value = temp2 
				k=k+1
	col = col+2
wb.save('Freqdat.xlsx')



